import tempfile
import unittest
from pathlib import Path

import openpyxl

from scene_excel_utils import (
    FRONT_ONLY_CAMERA_CONTRACT_LINE,
    normalize_scene_specifications_for_generation,
    read_unique_scenes_from_excel,
)


class TestSceneExcelUtilsNonCarla(unittest.TestCase):
    def test_normalizes_generic_five_camera_contract_to_front_only(self):
        spec = (
            "Road Context: urban arterial\n"
            "Camera Contract: Save synchronized front, front_left, front_right, rear, "
            "and drone_follow streams using matched frame ids.\n"
            "Event Contract: Something happens"
        )
        normalized = normalize_scene_specifications_for_generation(spec, scene_prompt="No camera preference")
        self.assertIn(FRONT_ONLY_CAMERA_CONTRACT_LINE, normalized)
        self.assertNotIn("front_left, front_right, rear, and drone_follow", normalized)

    def test_preserves_five_camera_contract_when_multi_angle_is_explicit(self):
        spec = (
            "Camera Contract: Save synchronized front, front_left, front_right, rear, "
            "and drone_follow streams using matched frame ids.\n"
            "Success Criteria: all views aligned"
        )
        prompt = "Generate multi-angle coverage with all 5 angles for analysis."
        normalized = normalize_scene_specifications_for_generation(spec, scene_prompt=prompt)
        self.assertIn("front_left, front_right, rear, and drone_follow", normalized)

    def test_reads_scene_specifications_from_named_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel = Path(tmp) / "scenes.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Master"
            ws.append(["Keyword", "Prompt", "Code Ouput", "Verification", "Scene Specifications"])
            ws.append(
                [
                    "Red Light Violation",
                    "Prompt A",
                    None,
                    None,
                    "Success criteria: keep lane clear and dense traffic visible.",
                ]
            )
            wb.save(excel)

            scenes = read_unique_scenes_from_excel(excel)
            self.assertEqual(len(scenes), 1)
            self.assertIn("dense traffic visible", scenes[0]["scene_specifications"])

    def test_fallback_uses_code_output_when_it_contains_success_criteria(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel = Path(tmp) / "scenes.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Master"
            ws.append(["Keyword", "Prompt", "Code Ouput"])
            ws.append(
                [
                    "Red Light Violation",
                    "Prompt A",
                    "Success criteria: at least 40 vehicles total and off-lane props.",
                ]
            )
            wb.save(excel)

            scenes = read_unique_scenes_from_excel(excel)
            self.assertEqual(len(scenes), 1)
            self.assertIn("at least 40 vehicles", scenes[0]["scene_specifications"])

    def test_deduplicates_keywords_across_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel = Path(tmp) / "scenes.xlsx"
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Master"
            ws1.append(["Keyword", "Prompt", "Scene Specifications"])
            ws1.append(["Red Light Violation", "Prompt A", "Spec A"])

            ws2 = wb.create_sheet("Secondary")
            ws2.append(["Keyword", "Prompt", "Scene Specifications"])
            ws2.append(["Red Light Violation", "Prompt B", "Spec B"])
            ws2.append(["Sudden Heavy Rain", "Prompt C", "Spec C"])
            wb.save(excel)

            scenes = read_unique_scenes_from_excel(excel)
            self.assertEqual(len(scenes), 2)
            self.assertEqual(scenes[0]["keyword"], "Red Light Violation")
            self.assertEqual(scenes[0]["prompt"], "Prompt A")
            self.assertEqual(scenes[1]["keyword"], "Sudden Heavy Rain")


if __name__ == "__main__":
    unittest.main()
