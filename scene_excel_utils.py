#!/usr/bin/env python3
"""Utilities for reading scene rows from the Excel source workbook."""

from __future__ import annotations

from pathlib import Path
import re
from typing import Any, Dict, List, Optional

import openpyxl


FRONT_ONLY_CAMERA_CONTRACT_LINE = (
    "Camera Contract: Save synchronized front stream only (single front camera) using matched frame ids."
)
_MULTI_CAMERA_VIEWS = ("front", "front_left", "front_right", "rear", "drone_follow")
_EXPLICIT_MULTI_ANGLE_HINTS = (
    "multi-angle",
    "multi angle",
    "5-angle",
    "five-angle",
    "5 angles",
    "five angles",
    "all views",
    "all five",
    "all 5",
)
_EXPLICIT_SINGLE_FRONT_HINTS = (
    "single front camera",
    "front stream only",
    "front-only",
    "front only",
)


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    compact = "".join(ch if ch.isalnum() else " " for ch in text)
    return " ".join(compact.split())


def _normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip())


def _camera_contract_line(line: str) -> bool:
    return _normalize_space(line).lower().startswith("camera contract:")


def _explicit_multi_angle_requested(scene_prompt: str, scene_specifications: str) -> bool:
    prompt_text = _normalize_space(scene_prompt).lower()
    spec_lines = [line for line in str(scene_specifications or "").splitlines() if not _camera_contract_line(line)]
    spec_text = _normalize_space(" ".join(spec_lines)).lower()
    haystack = f"{prompt_text} {spec_text}".strip()
    if any(h in haystack for h in _EXPLICIT_SINGLE_FRONT_HINTS):
        return False
    return any(h in haystack for h in _EXPLICIT_MULTI_ANGLE_HINTS)


def _is_legacy_generic_five_camera_contract(line: str) -> bool:
    lowered = _normalize_space(line).lower()
    if not lowered.startswith("camera contract:"):
        return False
    if not all(view in lowered for view in _MULTI_CAMERA_VIEWS):
        return False
    return "save synchronized" in lowered and "matched frame" in lowered


def normalize_scene_specifications_for_generation(
    scene_specifications: str,
    *,
    scene_prompt: str = "",
) -> str:
    """Normalize stale workbook defaults to the latest camera contract for generation paths.

    Workbook rows often still carry the old generic 5-camera contract. The latest scene-loop
    skill defaults to single front-camera capture unless multi-angle capture is explicitly requested.
    This normalization keeps explicit multi-angle requests intact while rewriting the legacy default.
    """

    raw = str(scene_specifications or "")
    if not raw.strip():
        return raw
    if _explicit_multi_angle_requested(scene_prompt, raw):
        return raw

    updated_lines: List[str] = []
    changed = False
    for line in raw.splitlines():
        if _is_legacy_generic_five_camera_contract(line):
            updated_lines.append(FRONT_ONLY_CAMERA_CONTRACT_LINE)
            changed = True
        else:
            updated_lines.append(line)
    if not changed:
        return raw
    return "\n".join(updated_lines)


def build_header_map(sheet) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col_idx in range(1, sheet.max_column + 1):
        key = normalize_header(sheet.cell(1, col_idx).value)
        if key and key not in headers:
            headers[key] = col_idx
    return headers


def _first_matching_col(header_map: Dict[str, int], names: List[str]) -> Optional[int]:
    for name in names:
        idx = header_map.get(normalize_header(name))
        if idx:
            return idx
    return None


def _scene_spec_from_row(
    sheet,
    row_idx: int,
    scene_spec_col: Optional[int],
    code_output_col: Optional[int],
) -> str:
    if scene_spec_col:
        raw = sheet.cell(row_idx, scene_spec_col).value
        if raw is not None and str(raw).strip():
            return str(raw).strip()

    # Backward-compat fallback: some files accidentally stored criteria in Code Ouput.
    if code_output_col:
        raw = sheet.cell(row_idx, code_output_col).value
        text = str(raw).strip() if raw is not None else ""
        lowered = text.lower()
        if text and (
            lowered.startswith("success criteria")
            or lowered.startswith("scene specifications")
            or lowered.startswith("scene specs")
        ):
            return text

    return ""


def read_unique_scenes_from_excel(excel_path: Path) -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(excel_path)
    scenes: List[Dict[str, Any]] = []
    seen_keywords = set()

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_map = build_header_map(sheet)

        keyword_col = _first_matching_col(
            header_map,
            ["Keyword", "Scene Keyword"],
        ) or 1
        prompt_col = _first_matching_col(
            header_map,
            ["Prompt", "Scene Prompt", "Description"],
        ) or 2
        scene_spec_col = _first_matching_col(
            header_map,
            [
                "Scene Specifications",
                "Scene Specification",
                "Scene Specs",
                "Success Criteria",
                "Scene Success Criteria",
            ],
        )
        code_output_col = _first_matching_col(
            header_map,
            ["Code Ouput", "Code Output"],
        )

        for row_idx in range(2, sheet.max_row + 1):
            keyword_raw = sheet.cell(row_idx, keyword_col).value
            prompt_raw = sheet.cell(row_idx, prompt_col).value
            if keyword_raw is None or prompt_raw is None:
                continue

            keyword = str(keyword_raw).strip()
            prompt = str(prompt_raw).strip()
            if not keyword or not prompt or keyword in seen_keywords:
                continue

            scene = {
                "serial": len(scenes) + 1,
                "keyword": keyword,
                "prompt": prompt,
                "scene_specifications": normalize_scene_specifications_for_generation(
                    _scene_spec_from_row(
                        sheet=sheet,
                        row_idx=row_idx,
                        scene_spec_col=scene_spec_col,
                        code_output_col=code_output_col,
                    ),
                    scene_prompt=prompt,
                ),
                "sheet": sheet_name,
                "row": row_idx,
            }
            scenes.append(scene)
            seen_keywords.add(keyword)

    return scenes
