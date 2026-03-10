#!/usr/bin/env python3
"""Build a concise per-shot history workbook from handoff + simulation logs."""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook


STATUS_PREPARED_RE = re.compile(r"status_update:\s+prepared_shot_(\d+)\s+for ready_for_wine at (.+)")
STATUS_READY_RE = re.compile(r"status_update:\s+ready_for_wine at (.+)")
SHOT_RE = re.compile(r"shot_(\d+)")
PHASE_RE = re.compile(r"crossing_phase=([A-Za-z]+)")


@dataclass
class ShotRecord:
    shot_index: int
    prepared_at: str = ""
    ready_at: str = ""
    code_file_posix: str = ""
    output_dir_posix: str = ""
    stage2_attempts: int = 0
    stage2_last_timestamp: str = ""
    stage2_last_success: str = ""
    stage2_last_returncode: str = ""
    stage2_last_frames_ok: str = ""
    stage2_last_coverage_ratio: str = ""
    stage2_last_total_frames: str = ""
    sim_log_path: str = ""
    sim_success: str = ""
    violation_logged: str = ""
    crossing_phases: str = ""
    dense_counts: str = ""
    nearby_visibility: str = ""
    notes: str = ""


def _extract_shot_index(text: str) -> Optional[int]:
    match = SHOT_RE.search(text)
    if not match:
        return None
    return int(match.group(1))


def _read_latest_manifest_path(handoff_dir: Path) -> Path:
    pointer = handoff_dir / "LATEST_MANIFEST_POSIX.txt"
    if not pointer.exists():
        raise FileNotFoundError(f"Missing latest manifest pointer: {pointer}")
    target = pointer.read_text(encoding="utf-8").strip()
    if not target:
        raise ValueError(f"Pointer file is empty: {pointer}")
    return Path(target)


def _load_manifest(manifest_path: Path) -> Dict[str, Any]:
    with manifest_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _parse_stage1_log(stage1_log: Path, records: Dict[int, ShotRecord]) -> None:
    if not stage1_log.exists():
        return

    current_prepared_shot: Optional[int] = None
    for raw_line in stage1_log.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line:
            continue

        prepared_match = STATUS_PREPARED_RE.match(line)
        if prepared_match:
            shot_index = int(prepared_match.group(1))
            timestamp = prepared_match.group(2).strip()
            rec = records.setdefault(shot_index, ShotRecord(shot_index=shot_index))
            rec.prepared_at = timestamp
            current_prepared_shot = shot_index
            continue

        ready_match = STATUS_READY_RE.match(line)
        if ready_match and current_prepared_shot is not None:
            rec = records.setdefault(current_prepared_shot, ShotRecord(shot_index=current_prepared_shot))
            rec.ready_at = ready_match.group(1).strip()
            current_prepared_shot = None
            continue

        if line.startswith("code_file_posix:"):
            if current_prepared_shot is None:
                guessed = _extract_shot_index(line)
                if guessed is None:
                    continue
                rec = records.setdefault(guessed, ShotRecord(shot_index=guessed))
                rec.code_file_posix = line.split(":", 1)[1].strip()
                continue
            rec = records.setdefault(current_prepared_shot, ShotRecord(shot_index=current_prepared_shot))
            rec.code_file_posix = line.split(":", 1)[1].strip()
        elif line.startswith("output_dir_posix:"):
            if current_prepared_shot is None:
                guessed = _extract_shot_index(line)
                if guessed is None:
                    continue
                rec = records.setdefault(guessed, ShotRecord(shot_index=guessed))
                rec.output_dir_posix = line.split(":", 1)[1].strip()
                continue
            rec = records.setdefault(current_prepared_shot, ShotRecord(shot_index=current_prepared_shot))
            rec.output_dir_posix = line.split(":", 1)[1].strip()


def _parse_stage2_logs(stage2_dir: Path, records: Dict[int, ShotRecord]) -> List[Dict[str, str]]:
    attempts: List[Dict[str, str]] = []
    stage2_logs = sorted(stage2_dir.glob("stage2_runner_*.log"))
    for log_path in stage2_logs:
        data: Dict[str, str] = {}
        lines = log_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        for line in lines:
            if ":" not in line:
                continue
            key, value = line.split(":", 1)
            data[key.strip()] = value.strip()

        shot_index = _extract_shot_index(data.get("code_file", "")) or _extract_shot_index(data.get("output_dir", ""))
        if shot_index is None:
            continue

        rec = records.setdefault(shot_index, ShotRecord(shot_index=shot_index))
        rec.stage2_attempts += 1

        timestamp = data.get("timestamp", "")
        if timestamp >= rec.stage2_last_timestamp:
            rec.stage2_last_timestamp = timestamp
            rec.stage2_last_success = data.get("success", "")
            rec.stage2_last_returncode = data.get("returncode", "")
            rec.stage2_last_frames_ok = data.get("frames_ok", "")
            rec.stage2_last_coverage_ratio = data.get("coverage_ratio", "")
            rec.stage2_last_total_frames = data.get("total_frames", "")
            if not rec.code_file_posix:
                rec.code_file_posix = data.get("code_file", "")
            if not rec.output_dir_posix:
                rec.output_dir_posix = data.get("output_dir", "")

        attempts.append(
            {
                "shot_index": str(shot_index),
                "timestamp": timestamp,
                "success": data.get("success", ""),
                "returncode": data.get("returncode", ""),
                "frames_ok": data.get("frames_ok", ""),
                "coverage_ratio": data.get("coverage_ratio", ""),
                "total_frames": data.get("total_frames", ""),
                "code_file": data.get("code_file", ""),
                "output_dir": data.get("output_dir", ""),
                "log_file": str(log_path.resolve()),
            }
        )

    return attempts


def _parse_sim_logs(scene_run_dir: Path, records: Dict[int, ShotRecord]) -> None:
    if not scene_run_dir.exists():
        return

    for shot_dir in sorted(scene_run_dir.glob("shot_*")):
        shot_index = _extract_shot_index(shot_dir.name)
        if shot_index is None:
            continue
        rec = records.setdefault(shot_index, ShotRecord(shot_index=shot_index))
        if not rec.output_dir_posix:
            rec.output_dir_posix = str(shot_dir.resolve())

        sim_logs = list(shot_dir.glob("*_simulation.log"))
        if not sim_logs:
            continue
        sim_log = max(sim_logs, key=lambda p: p.stat().st_mtime)
        rec.sim_log_path = str(sim_log.resolve())

        phases: Set[str] = set()
        has_success = False
        has_violation = False
        dense_line = ""
        nearby_line = ""
        has_error = False

        for raw_line in sim_log.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith("[SUCCESS]"):
                has_success = True
            if "[VIOLATION]" in line:
                has_violation = True
            if line.startswith("[ERROR]"):
                has_error = True
            if line.startswith("[INFO] Dense traffic counts:"):
                dense_line = line
            if line.startswith("[INFO] Nearby visibility traffic:"):
                nearby_line = line
            for match in PHASE_RE.finditer(line):
                phases.add(match.group(1))

        rec.sim_success = "yes" if has_success else "no"
        rec.violation_logged = "yes" if has_violation else "no"
        rec.crossing_phases = ", ".join(sorted(phases))
        rec.dense_counts = dense_line
        rec.nearby_visibility = nearby_line
        if has_error:
            rec.notes = "simulation log contains [ERROR]"


def _write_workbook(
    output_path: Path,
    run_id: str,
    scene_keyword: str,
    scenario_keyword: str,
    records: Dict[int, ShotRecord],
    attempts: List[Dict[str, str]],
) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "shots_summary"
    summary_headers = [
        "run_id",
        "scene_keyword",
        "scenario_keyword",
        "shot_index",
        "prepared_at",
        "ready_at",
        "code_file_posix",
        "output_dir_posix",
        "stage2_attempts",
        "stage2_last_timestamp",
        "stage2_last_success",
        "stage2_last_returncode",
        "stage2_last_frames_ok",
        "stage2_last_coverage_ratio",
        "stage2_last_total_frames",
        "sim_success",
        "violation_logged",
        "crossing_phases",
        "dense_counts",
        "nearby_visibility",
        "sim_log_path",
        "notes",
    ]
    ws.append(summary_headers)

    for shot_index in sorted(records):
        rec = records[shot_index]
        ws.append(
            [
                run_id,
                scene_keyword,
                scenario_keyword,
                rec.shot_index,
                rec.prepared_at,
                rec.ready_at,
                rec.code_file_posix,
                rec.output_dir_posix,
                rec.stage2_attempts,
                rec.stage2_last_timestamp,
                rec.stage2_last_success,
                rec.stage2_last_returncode,
                rec.stage2_last_frames_ok,
                rec.stage2_last_coverage_ratio,
                rec.stage2_last_total_frames,
                rec.sim_success,
                rec.violation_logged,
                rec.crossing_phases,
                rec.dense_counts,
                rec.nearby_visibility,
                rec.sim_log_path,
                rec.notes,
            ]
        )

    ws_attempts = wb.create_sheet("stage2_attempts")
    attempt_headers = [
        "shot_index",
        "timestamp",
        "success",
        "returncode",
        "frames_ok",
        "coverage_ratio",
        "total_frames",
        "code_file",
        "output_dir",
        "log_file",
    ]
    ws_attempts.append(attempt_headers)
    for item in sorted(attempts, key=lambda x: (int(x["shot_index"]), x["timestamp"])):
        ws_attempts.append([item[h] for h in attempt_headers])

    ws_meta = wb.create_sheet("meta")
    ws_meta.append(["field", "value"])
    ws_meta.append(["run_id", run_id])
    ws_meta.append(["scene_keyword", scene_keyword])
    ws_meta.append(["scenario_keyword", scenario_keyword])
    ws_meta.append(["total_shots", len(records)])
    ws_meta.append(["total_stage2_attempts", len(attempts)])
    ws_meta.append(["note", "Chat responses are in Codex conversation history, not auto-exported here."])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_history_workbook(manifest_path: Path, output_path: Optional[Path] = None) -> Path:
    manifest = _load_manifest(manifest_path)
    run_id = str(manifest["run_id"])
    scenario_keyword = str(manifest["scenario_keyword"])
    scene_keyword = str(manifest.get("scene_keyword", ""))

    base_dir = manifest_path.parents[3]
    handoff_run_dir = manifest_path.parent
    scene_run_dir = base_dir / "scenes" / scenario_keyword / run_id

    if output_path is None:
        output_path = handoff_run_dir / "shot_history.xlsx"
    else:
        output_path = output_path.resolve()

    records: Dict[int, ShotRecord] = {}
    stage1_log = handoff_run_dir / "stage1_agent_prepare.log"
    _parse_stage1_log(stage1_log, records)

    manifest_shot = manifest.get("shot_index")
    if isinstance(manifest_shot, int):
        records.setdefault(manifest_shot, ShotRecord(shot_index=manifest_shot))

    attempts = _parse_stage2_logs(handoff_run_dir, records)
    _parse_sim_logs(scene_run_dir, records)

    manifest_shot = manifest.get("shot_index")
    if isinstance(manifest_shot, int):
        rec = records.setdefault(manifest_shot, ShotRecord(shot_index=manifest_shot))
        if not rec.code_file_posix:
            rec.code_file_posix = manifest.get("code_file_posix", "")
        if not rec.output_dir_posix:
            rec.output_dir_posix = manifest.get("output_dir_posix", "")

    _write_workbook(
        output_path=output_path,
        run_id=run_id,
        scene_keyword=scene_keyword,
        scenario_keyword=scenario_keyword,
        records=records,
        attempts=attempts,
    )
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Build concise shot-history Excel from handoff logs")
    parser.add_argument("--manifest", help="Path to manifest.json (default: latest POSIX handoff)")
    parser.add_argument("--handoff-dir", default="handoffs", help="Handoff root for default manifest lookup")
    parser.add_argument("--output", help="Optional output .xlsx path")
    args = parser.parse_args()

    handoff_dir = Path(args.handoff_dir).resolve()
    manifest_path = Path(args.manifest).resolve() if args.manifest else _read_latest_manifest_path(handoff_dir)
    output_path = Path(args.output).resolve() if args.output else None
    generated = build_history_workbook(manifest_path=manifest_path, output_path=output_path)
    print(generated)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
