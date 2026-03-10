#!/usr/bin/env python3
"""Generate a single-sheet research history workbook (one row per scene run)."""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font

from scene_excel_utils import read_unique_scenes_from_excel


SHOT_RE = re.compile(r"shot_(\d+)")
PREPARED_SHOT_RE = re.compile(r"status_update:\s+prepared_shot_(\d+)\s+for ready_for_wine at (.+)")
READY_RE = re.compile(r"status_update:\s+ready_for_wine at (.+)")
FIX_REQ_RE = re.compile(r"SHOT_(\d+)\s+FIX REQUIREMENTS:\s*(.+)", re.IGNORECASE | re.DOTALL)
PHASE_RE = re.compile(r"crossing_phase=([A-Za-z]+)")


@dataclass
class ShotData:
    shot_index: int
    prepared_at: str = ""
    ready_at: str = ""
    code_file: str = ""
    output_dir: str = ""
    change_request: str = ""
    sim_success: bool = False
    violation_logged: bool = False
    map_fresh_loaded: bool = False
    dense_counts: str = ""
    nearby_visibility: str = ""
    crossing_phases: Set[str] = field(default_factory=set)
    sim_log_path: str = ""


@dataclass
class Stage2Attempt:
    shot_index: int
    timestamp: str
    success: bool
    process_success: bool
    frames_ok: bool
    coverage_ratio: Optional[float]
    total_frames: Optional[int]
    returncode: str
    note: str
    error: str
    code_file: str
    output_dir: str
    log_file: str
    is_dry_run: bool


def _extract_shot_index(text: str) -> Optional[int]:
    match = SHOT_RE.search(text or "")
    return int(match.group(1)) if match else None


def _parse_bool(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"true", "1", "yes", "y"}


def _parse_float(value: Any) -> Optional[float]:
    text = str(value or "").strip()
    if not text or text.lower() == "none":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_int(value: Any) -> Optional[int]:
    text = str(value or "").strip()
    if not text or text.lower() == "none":
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _safe_read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def _load_scene_meta(excel_path: Path) -> Dict[str, Dict[str, Any]]:
    if not excel_path.exists():
        return {}
    by_keyword: Dict[str, Dict[str, Any]] = {}
    for scene in read_unique_scenes_from_excel(excel_path):
        by_keyword[str(scene["keyword"]).strip().lower()] = scene
    return by_keyword


def _parse_stage1_log(stage1_log: Path, shots: Dict[int, ShotData]) -> Dict[str, str]:
    info: Dict[str, str] = {}
    text = _safe_read_text(stage1_log)
    if not text:
        return info

    lines = text.splitlines()
    current_prepared_shot: Optional[int] = None
    in_scene_prompt = False
    in_scene_specs = False
    prompt_lines: List[str] = []
    specs_lines: List[str] = []

    for raw in lines:
        line = raw.strip()
        if not line:
            if in_scene_prompt:
                in_scene_prompt = False
            if in_scene_specs:
                in_scene_specs = False
            continue

        if raw.startswith("scene_prompt_original:"):
            in_scene_prompt = True
            in_scene_specs = False
            continue

        if raw.startswith("scene_specifications:"):
            in_scene_specs = True
            in_scene_prompt = False
            continue

        if in_scene_prompt:
            prompt_lines.append(raw)
            continue

        if in_scene_specs:
            specs_lines.append(raw)
            continue

        prepared_match = PREPARED_SHOT_RE.match(line)
        if prepared_match:
            shot_idx = int(prepared_match.group(1))
            prepared_at = prepared_match.group(2).strip()
            shot = shots.setdefault(shot_idx, ShotData(shot_index=shot_idx))
            shot.prepared_at = prepared_at
            current_prepared_shot = shot_idx
            continue

        ready_match = READY_RE.match(line)
        if ready_match:
            if current_prepared_shot is not None:
                shot = shots.setdefault(current_prepared_shot, ShotData(shot_index=current_prepared_shot))
                shot.ready_at = ready_match.group(1).strip()
                current_prepared_shot = None
            continue

        if ":" in line:
            key, value = line.split(":", 1)
            key = key.strip()
            value = value.strip()
            if key in {
                "run_id",
                "scene_keyword",
                "scenario_keyword",
                "scene_serial",
                "generation_mode",
                "generation_status",
                "code_seed_source",
            }:
                info[key] = value

            if key == "code_file_posix":
                idx = _extract_shot_index(value)
                if idx is None and current_prepared_shot is not None:
                    idx = current_prepared_shot
                if idx is not None:
                    shot = shots.setdefault(idx, ShotData(shot_index=idx))
                    shot.code_file = value

            if key == "output_dir_posix":
                idx = _extract_shot_index(value)
                if idx is None and current_prepared_shot is not None:
                    idx = current_prepared_shot
                if idx is not None:
                    shot = shots.setdefault(idx, ShotData(shot_index=idx))
                    shot.output_dir = value

    if prompt_lines:
        info["scene_prompt_original"] = "\n".join(prompt_lines).strip()
    if specs_lines:
        info["scene_specifications"] = "\n".join(specs_lines).strip()
    return info


def _parse_stage2_log_file(path: Path) -> Dict[str, str]:
    data: Dict[str, str] = {}
    allowed = {
        "timestamp",
        "started_at",
        "ended_at",
        "manifest_path",
        "code_file",
        "script_path",
        "output_dir",
        "duration_seconds",
        "timeout_seconds",
        "script_args",
        "returncode",
        "process_success",
        "frames_ok",
        "success",
        "coverage_ratio",
        "total_frames",
        "best_stream_frames",
        "error",
    }
    note_lines: List[str] = []
    capture_note = False

    for raw in _safe_read_text(path).splitlines():
        line = raw.rstrip("\n")
        if capture_note:
            if not line.strip():
                capture_note = False
                continue
            note_lines.append(line.strip())
            continue

        if line.strip() == "note:":
            capture_note = True
            continue

        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = key.strip()
        value = value.strip()
        if key in allowed:
            data[key] = value

    if note_lines:
        data["note"] = " ".join(note_lines)
    return data


def _collect_stage2_attempts(scene_handoff_dir: Path) -> List[Stage2Attempt]:
    attempts: List[Stage2Attempt] = []
    for log_path in sorted(scene_handoff_dir.glob("stage2_runner_*.log")):
        data = _parse_stage2_log_file(log_path)
        code_path = data.get("code_file") or data.get("script_path") or ""
        output_dir = data.get("output_dir") or ""
        shot_idx = _extract_shot_index(code_path) or _extract_shot_index(output_dir)
        if shot_idx is None:
            continue

        attempts.append(
            Stage2Attempt(
                shot_index=shot_idx,
                timestamp=data.get("timestamp") or data.get("started_at") or "",
                success=_parse_bool(data.get("success")),
                process_success=_parse_bool(data.get("process_success")),
                frames_ok=_parse_bool(data.get("frames_ok")),
                coverage_ratio=_parse_float(data.get("coverage_ratio")),
                total_frames=_parse_int(data.get("total_frames")),
                returncode=data.get("returncode", ""),
                note=data.get("note", ""),
                error=data.get("error", ""),
                code_file=code_path,
                output_dir=output_dir,
                log_file=str(log_path.resolve()),
                is_dry_run="dry-run only" in str(data.get("note", "")).lower(),
            )
        )
    return attempts


def _extract_fix_requirement(text: str, shot_idx: int) -> str:
    if not text:
        return ""
    match = FIX_REQ_RE.search(text)
    if match and int(match.group(1)) == shot_idx:
        return " ".join(match.group(2).strip().split())

    marker = f"SHOT_{shot_idx} FIX REQUIREMENTS:"
    pos = text.upper().find(marker)
    if pos >= 0:
        after = text[pos + len(marker) :].strip()
        first_para = after.split("\n\n", 1)[0]
        return " ".join(first_para.split())
    return ""


def _collect_shot_prompt_changes(base_dir: Path, run_id: str, scenario_keyword: str, shots: Dict[int, ShotData]) -> None:
    prompt_root = base_dir / "generated_prompts" / scenario_keyword / run_id
    if not prompt_root.exists():
        return
    for shot_dir in sorted(prompt_root.glob("shot_*")):
        shot_idx = _extract_shot_index(shot_dir.name)
        if shot_idx is None:
            continue
        shot = shots.setdefault(shot_idx, ShotData(shot_index=shot_idx))
        enhanced = shot_dir / "enhanced_prompt.txt"
        text = _safe_read_text(enhanced)
        fix = _extract_fix_requirement(text, shot_idx)
        if fix:
            shot.change_request = fix


def _collect_sim_logs(base_dir: Path, run_id: str, scenario_keyword: str, shots: Dict[int, ShotData]) -> None:
    scene_root = base_dir / "scenes" / scenario_keyword / run_id
    if not scene_root.exists():
        return

    for shot_dir in sorted(scene_root.glob("shot_*")):
        shot_idx = _extract_shot_index(shot_dir.name)
        if shot_idx is None:
            continue
        shot = shots.setdefault(shot_idx, ShotData(shot_index=shot_idx))
        if not shot.output_dir:
            shot.output_dir = str(shot_dir.resolve())

        sim_logs = sorted(shot_dir.glob("*_simulation.log"), key=lambda p: p.stat().st_mtime)
        if not sim_logs:
            continue
        sim_log = sim_logs[-1]
        shot.sim_log_path = str(sim_log.resolve())
        text = _safe_read_text(sim_log)

        shot.sim_success = "[SUCCESS]" in text
        shot.violation_logged = "[VIOLATION]" in text
        shot.map_fresh_loaded = "Active map after fresh load" in text

        for line in text.splitlines():
            line_s = line.strip()
            if line_s.startswith("[INFO] Dense traffic counts:"):
                shot.dense_counts = line_s
            elif line_s.startswith("[INFO] Nearby visibility traffic:"):
                shot.nearby_visibility = line_s
            for match in PHASE_RE.finditer(line_s):
                shot.crossing_phases.add(match.group(1))


def _load_visual_eval(scene_handoff_dir: Path) -> Tuple[Optional[Path], Dict[str, Any]]:
    candidates = sorted(scene_handoff_dir.glob("*_visual_eval.json"), key=lambda p: p.stat().st_mtime)
    if not candidates:
        return None, {}
    path = candidates[-1]
    try:
        with path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        return path, data
    except Exception:
        return path, {}


def _shorten(text: str, limit: int = 600) -> str:
    if text is None:
        return ""
    txt = str(text).strip()
    if len(txt) <= limit:
        return txt
    return txt[: limit - 3] + "..."


def _join_unique(values: Iterable[str], sep: str = " | ") -> str:
    seen: Set[str] = set()
    ordered: List[str] = []
    for raw in values:
        if not raw:
            continue
        val = raw.strip()
        if not val or val in seen:
            continue
        seen.add(val)
        ordered.append(val)
    return sep.join(ordered)


def _build_scene_row(
    manifest_path: Path,
    manifest: Dict[str, Any],
    scene_meta: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    scene_handoff_dir = manifest_path.parent
    base_dir = manifest_path.parents[3]

    shots: Dict[int, ShotData] = {}
    stage1_log = scene_handoff_dir / "stage1_agent_prepare.log"
    stage1_info = _parse_stage1_log(stage1_log, shots)

    attempts = _collect_stage2_attempts(scene_handoff_dir)
    for attempt in attempts:
        shot = shots.setdefault(attempt.shot_index, ShotData(shot_index=attempt.shot_index))
        if not shot.code_file and attempt.code_file:
            shot.code_file = attempt.code_file
        if not shot.output_dir and attempt.output_dir:
            shot.output_dir = attempt.output_dir

    run_id = str(manifest.get("run_id", stage1_info.get("run_id", "")))
    scenario_keyword = str(manifest.get("scenario_keyword", stage1_info.get("scenario_keyword", "")))
    scene_keyword = str(manifest.get("scene_keyword", stage1_info.get("scene_keyword", "")))
    scene_serial = str(manifest.get("scene_serial", stage1_info.get("scene_serial", "")))
    generation_mode = str(manifest.get("generation_mode", stage1_info.get("generation_mode", "")))
    generation_status = str(manifest.get("generation_status", stage1_info.get("generation_status", "")))

    _collect_shot_prompt_changes(base_dir, run_id, scenario_keyword, shots)
    _collect_sim_logs(base_dir, run_id, scenario_keyword, shots)

    # Ensure final manifest shot exists.
    manifest_shot = manifest.get("shot_index")
    if isinstance(manifest_shot, int):
        shots.setdefault(manifest_shot, ShotData(shot_index=manifest_shot))

    eval_file, eval_data = _load_visual_eval(scene_handoff_dir)

    total_attempts = len(attempts)
    executed_attempts = [a for a in attempts if not a.is_dry_run]
    total_executed_attempts = len(executed_attempts)
    success_attempts = sum(1 for a in attempts if a.success)
    process_success_attempts = sum(1 for a in attempts if a.process_success)
    frames_ok_attempts = sum(1 for a in attempts if a.frames_ok)
    success_attempts_executed = sum(1 for a in executed_attempts if a.success)
    process_success_attempts_executed = sum(1 for a in executed_attempts if a.process_success)
    frames_ok_attempts_executed = sum(1 for a in executed_attempts if a.frames_ok)

    def rate(num: int, den: int) -> float:
        return round((num / den), 4) if den else 0.0

    coverage_values = [a.coverage_ratio for a in executed_attempts if a.coverage_ratio is not None]
    avg_coverage = round(sum(coverage_values) / len(coverage_values), 4) if coverage_values else 0.0

    latest_attempt = max(attempts, key=lambda a: a.timestamp) if attempts else None
    shot_indices = sorted(shots.keys())

    change_entries: List[str] = []
    status_entries: List[str] = []
    violation_shots: List[str] = []
    map_fresh_shots: List[str] = []
    phase_set: Set[str] = set()
    latest_dense = ""
    latest_nearby = ""
    latest_sim_log = ""

    for idx in shot_indices:
        shot = shots[idx]
        if shot.change_request:
            change_entries.append(f"shot_{idx}: {shot.change_request}")
        elif idx == 0:
            change_entries.append("shot_0: initial generation")
        else:
            change_entries.append(f"shot_{idx}: manual/code-only update (no explicit fix text)")

        phases = ",".join(sorted(shot.crossing_phases)) if shot.crossing_phases else "-"
        status_entries.append(
            f"shot_{idx}: sim_success={'yes' if shot.sim_success else 'no'}, "
            f"violation={'yes' if shot.violation_logged else 'no'}, phases={phases}"
        )
        if shot.violation_logged:
            violation_shots.append(f"shot_{idx}")
        if shot.map_fresh_loaded:
            map_fresh_shots.append(f"shot_{idx}")
        phase_set.update(shot.crossing_phases)
        if shot.dense_counts:
            latest_dense = shot.dense_counts
        if shot.nearby_visibility:
            latest_nearby = shot.nearby_visibility
        if shot.sim_log_path:
            latest_sim_log = shot.sim_log_path

    criteria_results = eval_data.get("criteria_results", []) if isinstance(eval_data, dict) else []
    criteria_names = [str(item.get("criterion", "")).strip() for item in criteria_results if isinstance(item, dict)]
    intent_checklist = eval_data.get("intent_checklist", []) if isinstance(eval_data, dict) else []

    scene_specs = str(manifest.get("scene_specifications", "") or stage1_info.get("scene_specifications", "")).strip()
    scene_prompt = str(manifest.get("scene_prompt_original", "") or stage1_info.get("scene_prompt_original", "")).strip()
    if not scene_specs and scene_keyword:
        lookup = scene_meta.get(scene_keyword.lower())
        if lookup:
            scene_specs = str(lookup.get("scene_specifications", "")).strip()
            if not scene_prompt:
                scene_prompt = str(lookup.get("prompt", "")).strip()

    eval_overall_pass = bool(eval_data.get("overall_pass")) if isinstance(eval_data, dict) else False
    human_verdict = str(eval_data.get("human_intent_verdict", "")).strip() if isinstance(eval_data, dict) else ""

    research_label = "PASS"
    if total_attempts == 0:
        research_label = "NO_SIM_RUN"
    elif success_attempts == 0:
        research_label = "SIM_FAIL"
    elif eval_data and not eval_overall_pass:
        research_label = "EVAL_FAIL"

    final_shot_index = manifest_shot if isinstance(manifest_shot, int) else (max(shot_indices) if shot_indices else 0)
    final_shot = shots.get(final_shot_index, ShotData(shot_index=final_shot_index))
    latest_stage2_success = bool(latest_attempt.success) if latest_attempt else False
    end_to_end_pass = latest_stage2_success and (eval_overall_pass if eval_data else True)

    row = {
        "run_id": run_id,
        "scene_serial": scene_serial,
        "scene_keyword": scene_keyword,
        "scenario_keyword": scenario_keyword,
        "scene_uid": f"{run_id}:{scenario_keyword}",
        "generation_mode": generation_mode,
        "generation_status": generation_status,
        "generated_at": str(manifest.get("generated_at", "")),
        "scene_prompt_original": _shorten(scene_prompt, 1000),
        "scene_specifications": _shorten(scene_specs, 1000),
        "total_shots": len(shot_indices),
        "shot_indices": ", ".join(f"shot_{i}" for i in shot_indices),
        "final_shot_index": final_shot_index,
        "final_code_file": final_shot.code_file or str(manifest.get("code_file_posix", "")),
        "final_output_dir": final_shot.output_dir or str(manifest.get("output_dir_posix", "")),
        "shot_change_history": _shorten(" || ".join(change_entries), 4000),
        "shot_status_history": _shorten(" || ".join(status_entries), 4000),
        "stage2_attempts_total": total_attempts,
        "stage2_attempts_executed": total_executed_attempts,
        "stage2_success_attempts": success_attempts,
        "stage2_success_attempts_executed": success_attempts_executed,
        "stage2_success_rate": rate(success_attempts_executed, total_executed_attempts),
        "stage2_success_rate_all": rate(success_attempts, total_attempts),
        "stage2_process_success_attempts": process_success_attempts,
        "stage2_process_success_attempts_executed": process_success_attempts_executed,
        "stage2_process_success_rate": rate(process_success_attempts_executed, total_executed_attempts),
        "stage2_process_success_rate_all": rate(process_success_attempts, total_attempts),
        "stage2_frames_ok_attempts": frames_ok_attempts,
        "stage2_frames_ok_attempts_executed": frames_ok_attempts_executed,
        "stage2_frames_ok_rate": rate(frames_ok_attempts_executed, total_executed_attempts),
        "stage2_frames_ok_rate_all": rate(frames_ok_attempts, total_attempts),
        "stage2_avg_coverage_ratio": avg_coverage,
        "stage2_last_timestamp": latest_attempt.timestamp if latest_attempt else "",
        "stage2_last_success": latest_attempt.success if latest_attempt else "",
        "stage2_last_returncode": latest_attempt.returncode if latest_attempt else "",
        "stage2_last_note": latest_attempt.note if latest_attempt else "",
        "stage2_last_error": latest_attempt.error if latest_attempt else "",
        "map_fresh_loaded_shots": ", ".join(map_fresh_shots),
        "violation_detected_shots": ", ".join(violation_shots),
        "crossing_phases_observed": ", ".join(sorted(phase_set)),
        "latest_dense_counts": latest_dense,
        "latest_nearby_visibility": latest_nearby,
        "visual_eval_exists": bool(eval_data),
        "visual_eval_mode": str(eval_data.get("evaluation_mode", "")) if isinstance(eval_data, dict) else "",
        "visual_eval_overall_pass": eval_overall_pass,
        "visual_eval_overall_pass_num": 1 if eval_overall_pass else 0,
        "visual_eval_human_verdict": human_verdict,
        "eval_criteria_passed": int(eval_data.get("criteria_passed", 0)) if isinstance(eval_data, dict) else 0,
        "eval_criteria_total": int(eval_data.get("criteria_total", 0)) if isinstance(eval_data, dict) else 0,
        "evaluation_criteria_asked": _join_unique(criteria_names, sep=" | "),
        "evaluation_intent_checklist": _join_unique((str(x) for x in intent_checklist), sep=" | "),
        "evaluation_summary": _shorten(str(eval_data.get("summary", "")), 1200) if isinstance(eval_data, dict) else "",
        "evaluation_suggested_fix": _shorten(str(eval_data.get("suggested_fix_prompt", "")), 1200)
        if isinstance(eval_data, dict)
        else "",
        "visual_eval_generated_at": str(eval_data.get("generated_at", "")) if isinstance(eval_data, dict) else "",
        "toolchain_end_to_end_pass": end_to_end_pass,
        "toolchain_end_to_end_pass_num": 1 if end_to_end_pass else 0,
        "research_ready_label": research_label,
        "stage1_log_file": str(stage1_log.resolve()) if stage1_log.exists() else "",
        "stage2_logs_count": len(list(scene_handoff_dir.glob("stage2_runner_*.log"))),
        "visual_eval_file": str(eval_file.resolve()) if eval_file else "",
        "latest_sim_log_file": latest_sim_log,
        "manifest_path": str(manifest_path.resolve()),
    }
    return row


def _sort_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def key(row: Dict[str, Any]) -> Tuple[str, str]:
        return (str(row.get("run_id", "")), str(row.get("scenario_keyword", "")))

    return sorted(rows, key=key)


def _autowidth(ws, max_width: int = 60) -> None:
    for col in ws.iter_cols(min_row=1, max_row=min(ws.max_row, 300)):
        letter = col[0].column_letter
        width = 10
        for cell in col:
            value = cell.value
            if value is None:
                continue
            width = max(width, len(str(value)))
        ws.column_dimensions[letter].width = min(max_width, width + 2)


def _add_charts(ws, first_data_row: int, last_data_row: int, columns: Dict[str, int], start_col: str = "AZ") -> None:
    if last_data_row < first_data_row:
        return

    # Chart 1: stage2 success rate
    chart1 = BarChart()
    chart1.title = "Stage2 Success Rate by Scene"
    chart1.y_axis.title = "Success Rate"
    chart1.x_axis.title = "Scene"
    data_ref = Reference(
        ws,
        min_col=columns["stage2_success_rate"],
        min_row=first_data_row - 1,
        max_row=last_data_row,
    )
    cat_ref = Reference(
        ws,
        min_col=columns["scene_uid"],
        min_row=first_data_row,
        max_row=last_data_row,
    )
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cat_ref)
    chart1.height = 8
    chart1.width = 18
    ws.add_chart(chart1, f"{start_col}2")

    # Chart 2: total shots per scene
    chart2 = BarChart()
    chart2.title = "Total Shots per Scene"
    chart2.y_axis.title = "Shots"
    chart2.x_axis.title = "Scene"
    data_ref2 = Reference(
        ws,
        min_col=columns["total_shots"],
        min_row=first_data_row - 1,
        max_row=last_data_row,
    )
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cat_ref)
    chart2.height = 8
    chart2.width = 18
    ws.add_chart(chart2, f"{start_col}22")

    # Chart 3: visual eval pass (binary)
    chart3 = BarChart()
    chart3.title = "Visual Eval Pass (1/0)"
    chart3.y_axis.title = "Pass"
    chart3.x_axis.title = "Scene"
    data_ref3 = Reference(
        ws,
        min_col=columns["visual_eval_overall_pass_num"],
        min_row=first_data_row - 1,
        max_row=last_data_row,
    )
    chart3.add_data(data_ref3, titles_from_data=True)
    chart3.set_categories(cat_ref)
    chart3.height = 8
    chart3.width = 18
    ws.add_chart(chart3, f"{start_col}42")


def build_research_scene_history(
    handoff_dir: Path,
    excel_source: Path,
    output_path: Path,
) -> Path:
    scene_meta = _load_scene_meta(excel_source)
    manifests = sorted(handoff_dir.glob("*/*/manifest.json"))
    rows: List[Dict[str, Any]] = []

    for manifest_path in manifests:
        try:
            manifest = json.loads(_safe_read_text(manifest_path))
            row = _build_scene_row(manifest_path=manifest_path, manifest=manifest, scene_meta=scene_meta)
            rows.append(row)
        except Exception as exc:
            rows.append(
                {
                    "run_id": "",
                    "scene_serial": "",
                    "scene_keyword": "",
                    "scenario_keyword": "",
                    "scene_uid": str(manifest_path),
                    "generation_mode": "",
                    "generation_status": "parse_error",
                    "generated_at": "",
                    "scene_prompt_original": "",
                    "scene_specifications": "",
                    "total_shots": 0,
                    "shot_indices": "",
                    "final_shot_index": "",
                    "final_code_file": "",
                    "final_output_dir": "",
                    "shot_change_history": "",
                    "shot_status_history": "",
                    "stage2_attempts_total": 0,
                    "stage2_success_attempts": 0,
                    "stage2_success_rate": 0.0,
                    "stage2_process_success_attempts": 0,
                    "stage2_process_success_rate": 0.0,
                    "stage2_frames_ok_attempts": 0,
                    "stage2_frames_ok_rate": 0.0,
                    "stage2_avg_coverage_ratio": 0.0,
                    "stage2_last_timestamp": "",
                    "stage2_last_success": "",
                    "stage2_last_returncode": "",
                    "stage2_last_note": "",
                    "stage2_last_error": str(exc),
                    "map_fresh_loaded_shots": "",
                    "violation_detected_shots": "",
                    "crossing_phases_observed": "",
                    "latest_dense_counts": "",
                    "latest_nearby_visibility": "",
                    "visual_eval_exists": False,
                    "visual_eval_mode": "",
                    "visual_eval_overall_pass": False,
                    "visual_eval_overall_pass_num": 0,
                    "visual_eval_human_verdict": "",
                    "eval_criteria_passed": 0,
                    "eval_criteria_total": 0,
                    "evaluation_criteria_asked": "",
                    "evaluation_intent_checklist": "",
                    "evaluation_summary": "",
                    "evaluation_suggested_fix": "",
                    "visual_eval_generated_at": "",
                    "research_ready_label": "PARSE_ERROR",
                    "stage1_log_file": "",
                    "stage2_logs_count": 0,
                    "visual_eval_file": "",
                    "latest_sim_log_file": "",
                    "manifest_path": str(manifest_path.resolve()),
                }
            )

    rows = _sort_rows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "scene_history"

    headers = [
        "run_id",
        "scene_serial",
        "scene_keyword",
        "scenario_keyword",
        "scene_uid",
        "generation_mode",
        "generation_status",
        "generated_at",
        "scene_prompt_original",
        "scene_specifications",
        "total_shots",
        "shot_indices",
        "final_shot_index",
        "final_code_file",
        "final_output_dir",
        "shot_change_history",
        "shot_status_history",
        "stage2_attempts_total",
        "stage2_attempts_executed",
        "stage2_success_attempts",
        "stage2_success_attempts_executed",
        "stage2_success_rate",
        "stage2_success_rate_all",
        "stage2_process_success_attempts",
        "stage2_process_success_attempts_executed",
        "stage2_process_success_rate",
        "stage2_process_success_rate_all",
        "stage2_frames_ok_attempts",
        "stage2_frames_ok_attempts_executed",
        "stage2_frames_ok_rate",
        "stage2_frames_ok_rate_all",
        "stage2_avg_coverage_ratio",
        "stage2_last_timestamp",
        "stage2_last_success",
        "stage2_last_returncode",
        "stage2_last_note",
        "stage2_last_error",
        "map_fresh_loaded_shots",
        "violation_detected_shots",
        "crossing_phases_observed",
        "latest_dense_counts",
        "latest_nearby_visibility",
        "visual_eval_exists",
        "visual_eval_mode",
        "visual_eval_overall_pass",
        "visual_eval_overall_pass_num",
        "visual_eval_human_verdict",
        "eval_criteria_passed",
        "eval_criteria_total",
        "evaluation_criteria_asked",
        "evaluation_intent_checklist",
        "evaluation_summary",
        "evaluation_suggested_fix",
        "visual_eval_generated_at",
        "toolchain_end_to_end_pass",
        "toolchain_end_to_end_pass_num",
        "research_ready_label",
        "stage1_log_file",
        "stage2_logs_count",
        "visual_eval_file",
        "latest_sim_log_file",
        "manifest_path",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"

    wrap_cols = {
        "scene_prompt_original",
        "scene_specifications",
        "shot_change_history",
        "shot_status_history",
        "evaluation_criteria_asked",
        "evaluation_intent_checklist",
        "evaluation_summary",
        "evaluation_suggested_fix",
        "latest_dense_counts",
        "latest_nearby_visibility",
        "stage2_last_note",
        "stage2_last_error",
    }
    col_map = {h: i + 1 for i, h in enumerate(headers)}

    for name in wrap_cols:
        col_idx = col_map[name]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    _autowidth(ws, max_width=70)
    _add_charts(ws, first_data_row=2, last_data_row=ws.max_row, columns=col_map, start_col="AZ")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Build single-sheet research scene history workbook")
    parser.add_argument("--handoff-dir", default="handoffs", help="Handoff root directory")
    parser.add_argument(
        "--excel-source",
        default="Keyword Prompt Verification.xlsx",
        help="Scene source workbook for prompt/spec fallback",
    )
    parser.add_argument(
        "--output",
        default="handoffs/research_scene_history.xlsx",
        help="Output xlsx path",
    )
    args = parser.parse_args()

    output = build_research_scene_history(
        handoff_dir=Path(args.handoff_dir).resolve(),
        excel_source=Path(args.excel_source).resolve(),
        output_path=Path(args.output).resolve(),
    )
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
